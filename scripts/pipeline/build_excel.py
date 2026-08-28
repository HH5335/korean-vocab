# -*- coding: utf-8 -*-
"""merged.json → data/韩语词表汇总.xlsx
- 汇总：每词一行：韩语、词性、中文释义、例句、例句来源（书/视频/AI/待AI）、例句中文
- 每本 PDF 书一张：同上 + 页码、书内序号
- AI待生成：例句来源=待AI 的行
- 来源冲突：同词多来源且释义不同的行
用法: .venv\\Scripts\\python build_excel.py
"""
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from pdf_common import DATA, EXTRACTED, MERGED_JSON, read_json_robust  # noqa: E402

OUT_XLSX = DATA / "韩语词表汇总.xlsx"
COLUMNS = ["韩语", "词性", "中文释义", "例句", "例句来源", "例句中文"]
FILL_YELLOW = "FFF3CD"  # 待AI
FILL_GREEN = "D9EDD7"   # 视频
HEADER_FILL = "F2E8F5"  # 蓝粉主题浅紫


def style_header(ws, ncols: int):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"


def set_widths(ws, widths: dict[int, int]):
    from openpyxl.utils import get_column_letter

    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def fill_row(ws, row: int, ncols: int, color: str):
    from openpyxl.styles import PatternFill

    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=color)


def main():
    from openpyxl import Workbook

    if not MERGED_JSON.exists():
        print("⚠️  缺少 merged.json，先运行 merge_vocab.py")
        return
    merged = read_json_robust(MERGED_JSON)
    conflicts = read_json_robust(EXTRACTED / "conflicts.json") if (EXTRACTED / "conflicts.json").exists() else []

    wb = Workbook()
    widths = {1: 22, 2: 10, 3: 34, 4: 56, 5: 10, 6: 40}

    # ---------- 汇总 ----------
    ws = wb.active
    ws.title = "汇总"
    ws.append(COLUMNS)
    rows_sorted = sorted(merged, key=lambda r: r["order"])
    for r in rows_sorted:
        ws.append([r["hangul"], r["pos"], r["meaning_cn"],
                   r["example_ko"], r["example_src"], r["example_zh"]])
    for i, r in enumerate(rows_sorted, start=2):
        if r["example_src"] == "待AI":
            fill_row(ws, i, len(COLUMNS), FILL_YELLOW)
        elif r["example_src"] == "视频":
            fill_row(ws, i, len(COLUMNS), FILL_GREEN)
    style_header(ws, len(COLUMNS))
    set_widths(ws, widths)

    # ---------- 每本 PDF 书一张 ----------
    books = {}
    for r in merged:
        if r["primary"].startswith("pdf:"):
            books.setdefault(r["primary"][4:], []).append(r)
    for book, recs in sorted(books.items()):
        if len(book) > 31:  # sheet 名上限 31 字符
            book = book[:31]
        ws = wb.create_sheet(title=book)
        cols = ["页码", "书内序号"] + COLUMNS
        ws.append(cols)
        for r in sorted(recs, key=lambda x: (int(x["ref"].get("page") or 0), int(x["ref"].get("seq") or 0))):
            ws.append([r["ref"].get("page", ""), r["ref"].get("seq", ""),
                       r["hangul"], r["pos"], r["meaning_cn"],
                       r["example_ko"], r["example_src"], r["example_zh"]])
            if r["example_src"] == "待AI":
                fill_row(ws, ws.max_row, len(cols), FILL_YELLOW)
        style_header(ws, len(cols))
        set_widths(ws, {1: 8, 2: 10, 3: 22, 4: 10, 5: 34, 6: 56, 7: 10, 8: 40})

    # ---------- AI待生成 ----------
    ws = wb.create_sheet(title="AI待生成")
    cols = ["韩语", "词性", "中文释义"]
    ws.append(cols)
    for r in rows_sorted:
        if r["example_src"] == "待AI":
            ws.append([r["hangul"], r["pos"], r["meaning_cn"]])
    style_header(ws, len(cols))
    set_widths(ws, {1: 22, 2: 10, 3: 34})

    # ---------- 来源冲突 ----------
    ws = wb.create_sheet(title="来源冲突")
    cols = ["韩语", "主来源", "主释义", "其他来源", "其他释义"]
    ws.append(cols)
    for c in conflicts:
        ws.append([c["hangul"], c["primary"], c["primary_meaning"],
                   c["other_source"], c["other_meaning"]])
    style_header(ws, len(cols))
    set_widths(ws, {1: 22, 2: 16, 3: 30, 4: 24, 5: 30})

    wb.save(OUT_XLSX)
    print(f"🎉 Excel 已生成 → {OUT_XLSX}")
    print(f"   汇总 {len(merged)} 行 | PDF 书 sheet {len(books)} 张 | 待AI {sum(1 for r in merged if r['example_src'] == '待AI')} 行 | 冲突 {len(conflicts)} 行")


if __name__ == "__main__":
    main()
